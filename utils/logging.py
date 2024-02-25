import json
import os

import openpyxl
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def save_results_to_excel(info_metrics, save_file_name):
    # Create a new workbook
    wb = openpyxl.Workbook()

    # Remove the default sheet created by openpyxl
    default_sheet = wb.active
    wb.remove(default_sheet)

    for sheet_name, info in info_metrics.items():
        ws = wb.create_sheet(title=sheet_name)
        depth = determine_info_depth(info)

        if depth == 2:
            populate_sheet_two_layers(info, ws)
        elif depth == 3:
            populate_sheet_three_layers(info, ws)
        else:
            print(f"Unknown depth ({depth}) for info structure in sheet '{sheet_name}'")

        # Adjust column widths
        adjust_column_widths(ws)

    # Save the workbook
    wb.save(filename=save_file_name)


def determine_info_depth(info):
    if isinstance(info, dict) and all(isinstance(v, dict) for v in info.values()):
        # Check if it's a two-layer or three-layer structure
        first_key = next(iter(info))
        if all(isinstance(v, dict) for v in info[first_key].values()):
            return 3  # Three-layer structure
        return 2  # Two-layer structure
    return None  # Unsupported structure


def populate_sheet_two_layers(info, ws):
    # Assuming the first key's value has all attributes
    first_round_info = info[next(iter(info))]
    headers = ['Round'] + list(first_round_info.keys())
    ws.append(headers)

    for round_idx, attributes in info.items():
        row = [round_idx] + [attributes.get(header) for header in headers[1:]]
        ws.append(row)


def populate_sheet_three_layers(info, ws):
    # Calculate max round number for all clients
    max_round = max(max(client_data.keys()) for client_data in info.values())
    # Initialize column for round headers
    round_col = 1
    ws.cell(row=2, column=round_col, value='Round')

    # Populate round numbers
    for round_num in range(0, max_round + 1):
        ws.cell(row=round_num + 3, column=round_col, value=round_num)

    start_col = 2  # Start from the second column for client data
    for client_id, client_data in sorted(info.items()):
        # Determine sub-headers and their respective spans
        sub_headers = {}
        for round_data in client_data.values():
            for key, value in round_data.items():
                if isinstance(value, list):
                    # Expand the sub-header for list-type values
                    for i in range(len(value)):
                        sub_headers[f"{key}_{i}"] = len(value)
                else:
                    sub_headers[key] = 1

        # Set the client header spanning all of its sub-headers
        client_header_col_end = start_col + sum(sub_headers.values()) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=client_header_col_end)
        ws.cell(row=1, column=start_col, value=f"Client {client_id}")
        ws.cell(row=1, column=start_col).alignment = Alignment(horizontal="center")

        # Write sub-headers
        sub_header_col = start_col
        for sub_header, span in sub_headers.items():
            if span > 1:
                # Merge cells for list-type sub-headers
                ws.merge_cells(start_row=2, start_column=sub_header_col, end_row=2, end_column=sub_header_col+span-1)
                ws.cell(row=2, column=sub_header_col, value=sub_header.split('_')[0]).alignment = Alignment(horizontal="center")
            else:
                ws.cell(row=2, column=sub_header_col, value=sub_header)
            sub_header_col += span

        for round_num in range(0, max_round + 1):
            row_num = round_num + 3  # Adjusting for header rows
            round_data = client_data.get(round_num, {})
            data_col = start_col  # Starting column for client data
            # Iterate directly over items in round_data
            for key, value in round_data.items():
                # Determine if the current item is part of an array
                if isinstance(value, list):
                    # Handle array: write each value in its own column
                    for i, val in enumerate(value):
                        ws.cell(row=row_num, column=data_col + i, value=val)
                    data_col += len(value)  # Move to the next column after the array
                else:
                    # Handle single value: write it and move to the next column
                    ws.cell(row=row_num, column=data_col, value=value)
                    data_col += 1  # Only increment by one for single values

        # Update the start column for the next client
        start_col = client_header_col_end + 1

    # Adjust column widths based on the longest entry in each column
    for col in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max_length + 2
def max_round(info):
    # Find the maximum round number across all clients
    return max(max(client_data.keys()) for client_data in info.values())

def adjust_column_widths(ws):
    for column in ws.columns:
        max_length = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column)
        adjusted_width = (max_length + 2)
        column_letter = get_column_letter(column[0].column)
        ws.column_dimensions[column_letter].width = adjusted_width



def json_str_to_int_key_dict(json_str):
    """
    将 JSON 字符串转换为字典，同时将字典的主键从字符串转换为整数。

    :param json_str: JSON 字符串
    :return: 转换后的字典
    """
    # 解析 JSON 字符串
    original_dict = json.loads(json_str)
    # 将所有键转换为整数
    converted_dict = {int(k): v for k, v in original_dict.items()}
    return converted_dict
