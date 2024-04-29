import base64
import colorsys
import io
import os
import random
import shutil
import time
from datetime import datetime

import aiohttp
import pandas as pd
import requests
from ex4nicegui.reactive import local_file_picker, rxui
from ex4nicegui.utils.signals import to_ref_wrapper, to_ref, on
from nicegui import ui, context
from openpyxl.styles import Alignment
from openpyxl.workbook import Workbook

from visual.parts.constant import record_names, record_types, type_name_mapping, user_info_mapping
from visual.parts.local_file_picker import local_file_picker


def algo_to_sheets(config):
    # 初始化存放所有sheets数据的字典
    sheets_dict = {}

    # 处理除algo_params之外的其他配置，这些会放在一个单独的sheet中
    main_config = {k: v for k, v in config.items() if k != 'algo_params'}
    sheets_dict['main_config'] = pd.DataFrame([main_config])

    # 处理algo_params配置，每个元素创建一个新的sheet
    for algo_param in config['algo_params']:
        # 参数id作为sheet的名字
        sheet_name = f"algo_params_{algo_param['id']}"
        # 分离'params'和其他配置
        param_config = algo_param.pop('params')
        other_config = algo_param  # 包括id, type, spot, algo

        # 拆分params中的数组，使每个值成为一行
        max_len = max(len(v) if isinstance(v, list) else 1 for v in param_config.values())
        expanded_params = []
        for i in range(max_len):
            row = {k: v[i] if isinstance(v, list) and i < len(v) else v for k, v in param_config.items()}
            expanded_params.append(row)

        # 将拆分后的参数与其他配置合并
        full_config_rows = []
        for row in expanded_params:
            full_row = {**other_config, **row}
            full_config_rows.append(full_row)

        # 将配置转换为DataFrame
        sheet_data = pd.DataFrame(full_config_rows)
        sheets_dict[sheet_name] = sheet_data

    return sheets_dict

def ext_info(info_dict, target_key):
    # 递归函数来展开指定键
    def recurse_extract(sub_dict, target_key):
        if target_key in sub_dict:  # 如果找到目标键
            # 返回目标键对应的值
            return sub_dict[target_key]
        else:  # 否则递归查找每个字典类型的值
            flag = False
            res = {}
            for k, v in sub_dict.items():
                if isinstance(v, dict):
                    result = recurse_extract(v, target_key)
                    if result is not None:
                        flag=True
                        res[k] = result
            if flag:
                return res
            else:
                return None

    # 开始处理传入的字典
    new_dict = {}
    for k, v in info_dict.items():
        if k == target_key:
            new_dict = v  # 直接取值，不再递归
        elif isinstance(v, dict):
            # 递归查找并构建新的字典
            result = recurse_extract(v, target_key)
            if result is not None:  # 找到了目标键
                new_dict[k] = result
            else:  # 未找到目标键，保持原结构
                new_dict[k] = v
        else:  # 非字典类型的值，直接保留
            new_dict[k] = v

    return new_dict




# 转换键的辅助函数
def convert_keys_to_int(data):
    if isinstance(data, dict):
        new_dict = {}
        for key, value in data.items():
            if key.isdigit():  # 检查键是否为数字字符串
                int_key = int(key)  # 转换为整数类型
            else:
                int_key = key  # 保持为原始类型
            new_dict[int_key] = convert_keys_to_int(value)  # 递归调用
        return new_dict
    elif isinstance(data, list):
        return [convert_keys_to_int(item) for item in data]  # 处理列表中的每个元素
    else:
        return data  # 非容器类型直接返回原值


def get_image_data(image_path):
    with open(image_path, 'rb') as image_file:
        image_data = image_file.read()
    return image_data


def to_base64(input_data):
    """
    将给定的在线图片URL或本地图片的二进制内容转换为Base64编码。
    如果输入是URL，则从该URL下载图片并转换为Base64编码。
    如果输入是二进制内容，则直接转换为Base64编码。

    :param input_data: 图片的URL字符串或图片的二进制内容。
    :return: 图片的Base64编码字符串。
    """
    # 检查是否是URL
    if isinstance(input_data, str) and input_data.startswith('http'):
        # 从URL下载图片
        response = requests.get(input_data)
        if response.status_code == 200:
            # 将图片内容转换为Base64编码
            image_data = response.content
        else:
            raise ValueError("无法从URL下载图片")
    elif isinstance(input_data, bytes):
        # 如果输入是二进制内容，直接使用
        image_data = input_data
    else:
        raise ValueError("输入类型不支持，必须是图片URL或二进制内容")
    # 将图片内容转换为Base64编码
    return 'data:image/png;base64,' + base64.b64encode(image_data).decode('utf-8')


async def han_fold_choice(ref):
    result = await local_file_picker(ref.value, upper_limit=None, directories_only=True)
    if result is not None and len(result) != 0:
        ref.value = result[0]


async def han_file_choice(ref, limited):
    result = await local_file_picker(ref.value, upper_limit=None, multiple=True, allowed_file_types=limited)
    # 假设result返回的是选中文件的路径列表
    if result:
        for item in result:
            if item not in ref.value:
                ref.value.append(item)  # 若不存在，则添加到ref.value中


def my_vmodel(data, key):
    def setter(new):
        data[key] = new
    return to_ref_wrapper(lambda: data[key], setter)


def convert_dict_to_list(mapping, mapping_dict):
    mapping_list = []
    for key, value in mapping.items():
        in_dict = {'id': key}
        if type(value) is not list:
            value = (value,)
        for i, k in enumerate(mapping_dict):
            in_dict[k] = value[i]
        mapping_list.append(in_dict)
    return mapping_list


def convert_list_to_dict(mapping_list, mapping_dict):
    mapping = {}
    for item in mapping_list:
        if len(mapping_dict) == 1:
            for k in mapping_dict:
                mapping[item['id']] = item[k]
        else:
            in_list = []
            for k in mapping_dict:
                in_list.append(item[k])
            mapping[item['id']] = in_list[0] if len(in_list) == 1 else in_list
    return mapping


def convert_tuple_to_dict(mapping, mapping_dict):
    new_dict = {}
    for i, k in enumerate(mapping_dict):
        new_dict[k] = mapping[i]
    return new_dict


def convert_dict_to_tuple(mapping):
    new_list = []
    for k, v in mapping.items():
        new_list.append(v)
    return tuple(new_list)


def get_dicts(colors, infos_each, infos_all, ):
    legend_dict = [
        {
            'name': '总数',
            'icon': 'circle',
        },
        {
            'name': '噪声',
            'icon': 'circle',
        }
    ]
    for label in infos_each:
        legend_dict.append(
            {
                'name': '类别' + str(label),
                'icon': 'circle',
            }
        )
    series_dict = [
        {
            "data": infos_all['total'],
            "type": "bar",
            "name": '总数',
            "barWidth": '10%',  # 设置柱子的宽度
            "barCategoryGap": '0%',  # 设置类目间柱形距离（类目宽的百分比）
            'itemStyle': {
                'color': 'black',
            },
            'emphasis': {
                'focus': 'self',
                'itemStyle': {
                    'borderColor': 'black',
                    'color': 'black',
                    'borderWidth': 20,
                    'shadowBlur': 10,
                    'shadowOffsetX': 0,
                    'shadowColor': 'rgba(0, 0, 0, 0)',
                    'scale': True  # 实际放大柱状图的部分
                },
                'label': {
                    'show': True,
                    'formatter': '{b}\n数量{c}',
                    'position': 'inside',
                }
            },
            # 如果您想要放大当前柱子
            'label': {
                'show': False
            },
            'labelLine': {
                'show': False
            }
        },
        {
            "data": infos_all['noise'],
            "type": "bar",
            "name": '噪声',
            "barWidth": '10%',  # 设置柱子的宽度
            "barCategoryGap": '20%',  # 设置类目间柱形距离（类目宽的百分比）
            'itemStyle': {
                'color': 'grey',
            },
            'emphasis': {
                'focus': 'self',
                'itemStyle': {
                    'borderColor': 'grey',
                    'color': 'grey',
                    'borderWidth': 20,
                    'shadowBlur': 10,
                    'shadowOffsetX': 0,
                    'shadowColor': 'rgba(0, 0, 0, 0)',
                    'scale': True  # 实际放大柱状图的部分
                },
                'label': {
                    'show': True,
                    'formatter': '{b}\n噪声{c}',
                    'position': 'inside',
                }
            },
            # 如果您想要放大当前柱子
            'label': {
                'show': False
            },
            'labelLine': {
                'show': False
            }
        }
    ]
    for label, (distribution, noises) in infos_each.items():
        series_dict.append(
            {
                "data": [
                    {
                        "value": dist,  # 总数据量
                        "itemStyle": {
                            "color": {
                                "type": "linear",
                                "x": 0,
                                "y": 0,
                                "x2": 0,
                                "y2": 1,
                                "colorStops": [
                                    {"offset": 0, "color": colors[label]},  # 原始颜色
                                    {"offset": (1 - noise / dist) if dist != 0 else 0, "color": colors[label]},
                                    # 与原始颜色相同，此处为噪声数据位置
                                    {"offset": (1 - noise / dist) if dist != 0 else 1, "color": 'grey'},  # 从噪声数据位置开始渐变
                                    {"offset": 1, "color": 'grey'}  # 底部透明
                                ]
                            }
                        }
                    }
                    for dist, noise in zip(distribution, noises)
                ],
                "type": "bar",
                "stack": 'each',
                "name": '类别' + str(label),
                "barWidth": '10%',  # 设置柱子的宽度
                "barCategoryGap": '20%',  # 设置类目间柱形距离（类目宽的百分比）
                'itemStyle': {
                    'color': colors[label],
                },
                'emphasis': {
                    'focus': 'self',
                    'itemStyle': {
                        'borderColor': colors[label],
                        'color': colors[label],
                        'borderWidth': 20,
                        'shadowBlur': 10,
                        'shadowOffsetX': 0,
                        'shadowColor': 'rgba(0, 0, 0, 0)',
                        'scale': True  # 实际放大柱状图的部分
                    },
                    'label': {
                        'show': True,
                        'formatter': '{b}\n数量{c}',
                        'position': 'inside',
                    }
                },
                # 如果您想要放大当前柱子
                'label': {
                    'show': False
                },
                'labelLine': {
                    'show': False
                }
            })
    return legend_dict, series_dict


def build_task_loading(message: str, is_done=False, state='positive'):
    with ui.row().classes("flex-center"):
        if not is_done:
            ui.spinner(color="negative")
        else:
            ui.icon("done", color=state)
        with ui.row():
            ui.label(message)


def get_n_hls_colors(num):
    hls_colors = []
    i = 0
    step = 360.0 / num
    while i < 360:
        h = i
        s = 90 + random.random() * 10
        l = 50 + random.random() * 10
        _hlsc = [h / 360.0, l / 100.0, s / 100.0]
        hls_colors.append(_hlsc)
        i += step

    return hls_colors


def ncolors(num):
    rgb_colors = []
    if num < 1:
        return rgb_colors
    hls_colors = get_n_hls_colors(num)
    for hlsc in hls_colors:
        _r, _g, _b = colorsys.hls_to_rgb(hlsc[0], hlsc[1], hlsc[2])
        r, g, b = [int(x * 255.0) for x in (_r, _g, _b)]
        rgb_colors.append([r, g, b])
    return rgb_colors


def color(value):
    digit = list(map(str, range(10))) + list("ABCDEF")
    if isinstance(value, tuple):
        string = '#'
        for i in value:
            a1 = i // 16
            a2 = i % 16
            string += digit[a1] + digit[a2]
        return string
    elif isinstance(value, str):
        a1 = digit.index(value[1]) * 16 + digit.index(value[2])
        a2 = digit.index(value[3]) * 16 + digit.index(value[4])
        a3 = digit.index(value[5]) * 16 + digit.index(value[6])
        return (a1, a2, a3)


def cal_dis_dict(infos, target='训练集'):
    print(infos)
    infos_each = infos['each']
    infos_all = infos['all']
    num_clients = 0 if len(infos_each) == 0 else len(infos_each[0][0])  # 现在还有噪声数据，必须取元组的首元素
    num_classes = len(infos_each)
    colors = list(map(lambda x: color(tuple(x)), ncolors(num_classes)))
    legend_dict, series_dict = get_dicts(colors, infos_each, infos_all)
    return {
        "toolbox": {
            "show": True,
            "feature": {
                "saveAsImage": {
                    "show": True,
                    "title": "下载图像",
                    "type": "png",
                    "lang": ["点击右键 -> 保存图片"]  # 提示用户如何保存图片
                }
            }
        },
        "xAxis": {
            "type": "category",
            "name": target + 'ID',
            "data": [target + str(i) for i in range(num_clients)],
        },
        "yAxis": {
            "type": "value",
            "name": '样本分布',
            "minInterval": 1,  # 设置Y轴的最小间隔
            "axisLabel": {
                'interval': 'auto',  # 根据图表的大小自动计算步长
            },
        },
        'legend': {
            'data': legend_dict,
            'type': 'scroll',  # 启用图例的滚动条
            'pageButtonItemGap': 5,
            'pageButtonGap': 20,
            'pageButtonPosition': 'end',  # 将翻页按钮放在最后
        },
        "series": series_dict,
        'tooltip': {
            'trigger': 'item',
            'axisPointer': {
                'type': 'shadow'
            },
            'formatter': "{b} <br/>{a} <br/> 数量{c}",
            'extraCssText': 'box-shadow: 0 0 8px rgba(0, 0, 0, 0.3);'  # 添加阴影效果
        },
        'grid': {
            'left': '3%',
            'right': '4%',
            'bottom': '10%',
            'containLabel': True
        },
        'dataZoom': [{
            'type': 'slider',
            'xAxisIndex': [0],
            'start': 10,
            'end': 90,
            'height': 5,
            'bottom': 10,
            # 'showDetail': False,
            'handleIcon': 'M8.2,13.4V6.2h4V2.2H5.4V6.2h4v7.2H5.4v4h7.2v-4H8.2z',
            'handleSize': '80%',
            'handleStyle': {
                'color': '#fff',
                'shadowBlur': 3,
                'shadowColor': 'rgba(0, 0, 0, 0.6)',
                'shadowOffsetX': 2,
                'shadowOffsetY': 2
            },
            'textStyle': {
                'color': "transparent"
            },
            # 使用 borderColor 透明来隐藏非激活状态的边框
            'borderColor': "transparent"
        }],
    }


def get_grad_info(info_ref):
    with ui.row().classes('w-full justify-center') as row:
        ui.label('请先执行任务')
    def han_update():
        row.clear()
        with row:
            if len(info_ref.value) != 0:
                for cid, r_per in info_ref.value[-1].items():
                    with ui.column():
                        ui.icon('cloud', color='blue').classes('text-5xl').tooltip(f'奖励梯度的全局占比{100 * r_per:.2f}%').style(f'opacity: {r_per}')
                        ui.label(f'客户{cid}').classes('w-full')
            else:
                ui.label('请等待任务执行...')
    on(info_ref)(han_update)


def get_user_info(info_ref, info_name, task_name):
    with ui.row().classes('w-full justify-center') as row:
        ui.label('请先执行任务')
    columns = [
        {"name": "round", "label": "轮次", "field": "id", 'align': 'center'},
        {"name": "bid", "label": "声明成本", "field": "bid", 'align': 'center'},
        {"name": "cost", "label": "真实成本", "field": "cost", 'align': 'center'},
        {"name": "score", "label": "得分", "field": "score", 'align': 'center'},
        {"name": "emp", "label": "经验指标", "field": "emp", 'align': 'center'},
        {"name": "ucb", "label": "UCB指标", "field": "ucb", 'align': 'center'},
        {"name": "idx", "label": "选择指标", "field": "idx", 'align': 'center'},
        {"name": "pay", "label": "支付", "field": "pay", 'align': 'center'},
        {"name": "util", "label": "效用", "field": "util", 'align': 'center'},
        {"name": "contrib", "label": "贡献", "field": "contrib", 'align': 'center'},
        {"name": "reward", "label": "奖励", "field": "reward", 'align': 'center'},
        {"name": "times", "label": "累计选中次数", "field": "times", 'align': 'center'},
    ]
    def download_data():
        # print(his_info)
        wb = Workbook()
        wb.remove(wb.active)
        for cid, infos in his_info.items():
            ws = wb.create_sheet(title='客户'+str(cid))
            # 初始化行和列索引
            row_index = 1  # 第一行为任务名称，第二行开始为数据类型和数据
            column_index = 1
            num = len(infos)
            for item in columns:
                ws.cell(row=1, column=column_index).value = item['label']
                for i in range(num):
                    ws.cell(row=row_index + 1 + i, column=column_index).value = infos[i][item['name']]
                column_index += 1
        data = io.BytesIO()
        wb.save(data)
        data.seek(0)
        ui.download(data.read(), get_local_download_path(info_name, task_name))

    his_info, icon_list, dialogs, tables = {}, {}, {}, {}  # 此容器用于存放历史的信息记录


    def han_update():
        if len(info_ref.value) != 0:
            if info_ref.value[-1][0] == 'statue':  # 若是状态的更新
                if len(icon_list) == 0:
                    row.clear()
                    with row:
                        ui.button('下载数据', icon='download', on_click=download_data).props('icon=cloud_download')
                        for cid, statue in info_ref.value[-1][1].items():
                            if cid not in his_info:
                                his_info[cid] = []  # 每个客户信息用列表存放绑定table
                            with ui.dialog() as dialogs[cid], ui.card().classes('w-full'):
                                tables[cid] = ui.table(columns=columns, rows=his_info[cid]).classes('w-full')
                            with ui.column():
                                lay = '落选' if statue == 'gray' else '获胜'
                                icon_list[cid] = ui.icon('person', color=statue).classes('text-5xl').on('click',
                                                                                                        lambda cid=cid:
                                                                                                        dialogs[cid].open()).tooltip(lay)
                                ui.label(f'客户{cid}').classes('w-full')
                else:
                    for cid, statue in info_ref.value[-1][1].items():
                        icon_list[cid].props('color=gray').tooltip('落选' if statue == 'gray' else '获胜')

            elif info_ref.value[-1][0] == 'info':  # 若是信息的更新
                this_info = info_ref.value[-1][1]
                for cid, info in this_info.items():  # 更新每个用户的基本信息
                    if cid not in his_info:
                        his_info[cid] = []
                    his_info[cid].append(info)
                    tables[cid].update()
    on(info_ref)(han_update)

# 全局信息使用算法-指标-轮次/时间的方式展示
def get_global_option(infos_dict, mode_ref, info_name, task_names):
    record_type = record_types['global']['param'][info_name]
    record_name = record_names['global']['param'][info_name]
    li = record_type.split('_')
    record_type = li[0]
    alloc_type = li[1] if len(li) > 1 else ''
    new_series, new_names, new_x = [], [], None
    if alloc_type == 'bul' and len(infos_dict[mode_ref.value][0].value) != 0:
        new_names = list(list(infos_dict[mode_ref.value][0].value)[-1][1].keys())
        if record_type == 'line':  # 需要按x轴值排序
            paired_data = [
                (item[0], {name: item[1][name] for name in new_names})
                for item in infos_dict[mode_ref.value][0].value
            ]
            paired_data.sort(key=lambda x: x[0])
        else:
            paired_data = infos_dict[mode_ref.value][0].value
        for name in new_names:
            data = [(item[0], item[1][name]) for item in paired_data]
            new_series.append({
                'name': name,
                'data': data,
                'type': record_type
            })

    options = {
        "toolbox": {
            "show": True,
            "feature": {
                "saveAsImage": {
                    "show": True,
                    "title": "下载图像",
                    "type": "png",
                    "lang": ["点击右键 -> 保存图片"]  # 提示用户如何保存图片
                }
            }
        },
        'grid': {
            "left": '10%',
            "top": '10%',
            "right": '10%',
            "bottom": '10%',
            "containLabel": True
        },
        'tooltip': {
            'trigger': 'axis',
            'axisPointer': {
                'type': 'cross',
                'lineStyle': {  # 设置纵向指示线
                    'type': 'dashed',
                    'color': "rgba(198, 196, 196, 0.75)"
                }
            },
            'crossStyle': {  # 设置横向指示线
                'color': "rgba(198, 196, 196, 0.75)"
            },
            'formatter': "算法{a}<br/>" + record_names['global']['type'][
                mode_ref.value] + ',' + record_type + "<br/>{c}",
            'extraCssText': 'box-shadow: 0 0 8px rgba(0, 0, 0, 0.3);'  # 添加阴影效果
        },
        "xAxis": {
            "type": 'category' if record_type == 'bar' else 'value',
            "name": record_names['global']['type'][mode_ref.value]
            if info_name not in type_name_mapping else type_name_mapping[info_name],
            'minInterval': 1 if mode_ref.value == 'round' else None,
        },
        "yAxis": {
            "type": "value",
            "name": record_name,
            "axisLabel": {
                'interval': 'auto',  # 根据图表的大小自动计算步长
            },
            'axisLine': {
                'show': True
            },
            'splitNumber': 5,  # 分成5个区间
        },
        'legend': {
            'data': [task_names[tid] for tid in task_names] if alloc_type != 'bul' or
                                                               len(infos_dict[mode_ref.value][
                                                                       0].value) == 0 else new_names,
            'type': 'scroll',  # 启用图例的滚动条
            'orient': 'horizontal',  # 横向排列
            'pageButtonItemGap': 5,
            'pageButtonGap': 20,
            'pageButtonPosition': 'end',  # 将翻页按钮放在最后
            'textStyle': {
                'overflow': 'truncate',  # 当文本超出宽度时，截断文本
                'ellipsis': '...',  # 截断时末尾添加的字符串
            },
            'tooltip': {
                'show': True  # 启用悬停时的提示框
            }
        },
        'series': [
            {
                'name': task_names[tid],
                'type': record_type,
                'data': list(infos_dict[mode_ref.value][tid].value),  # 受制于rx，这里必须告知前端为list类型
                'connectNulls': True,  # 连接数据中的空值
            }
            for tid in infos_dict[mode_ref.value]
        ] if alloc_type != 'bul' or len(infos_dict[mode_ref.value][0].value) == 0 else new_series,
        'dataZoom': [
            {
                'type': 'inside',  # 放大和缩小
                'orient': 'vertical',
                'start': 0,
                'end': 100,
                'minSpan': 1,  # 最小缩放比例，可以根据需要调整
                'maxSpan': 100,  # 最大缩放比例，可以根据需要调整
            },
            {
                'type': 'inside',
                'start': 0,
                'end': 100,
                'minSpan': 1,  # 最小缩放比例，可以根据需要调整
                'maxSpan': 100,  # 最大缩放比例，可以根据需要调整
            }
        ],
    }
    if alloc_type == 'seg' and list(infos_dict[mode_ref.value][0].value):
        m = max(max(t) for t in list(infos_dict[mode_ref.value][0].value))
        options['series'].append({
            "type": 'line',
            "data": [[0, 0], [m, m]],  # y=x
            "lineStyle": {
                "color": 'red',
                "type": 'dashed',
                "width": 1
            }
        })
    return options


# 局部信息使用客户-指标-轮次的方式展示，暂不支持算法-时间的显示
def get_local_option(info_dict: dict, mode_ref, info_name: str):
    return {
        "toolbox": {
            "show": True,
            "feature": {
                "saveAsImage": {
                    "show": True,
                    "title": "下载图像",
                    "type": "png",
                    "lang": ["点击右键 -> 保存图片"]  # 提示用户如何保存图片
                }
            }
        },
        'grid': {
            "left": '10%',
            "top": '10%',
            "right": '10%',
            "bottom": '10%',
            "containLabel": True
        },
        'tooltip': {
            'trigger': 'axis',
            'axisPointer': {
                'type': 'cross',
                'lineStyle': {  # 设置纵向指示线
                    'type': 'dashed',
                    'color': "rgba(198, 196, 196, 0.75)"
                }
            },
            'crossStyle': {  # 设置横向指示线
                'color': "rgba(198, 196, 196, 0.75)"
            },
            'formatter': "客户{a}<br/>" + record_names['local']['type'][mode_ref.value] + ',' +
                         record_names['local']['param'][info_name] + "<br/>{c}",
            'extraCssText': 'box-shadow: 0 0 8px rgba(0, 0, 0, 0.3);'  # 添加阴影效果
        },
        "xAxis": {
            "type": 'value',
            "name": record_names['local']['type'][mode_ref.value],
            'minInterval': 1 if mode_ref.value == 'round' else None,
        },
        "yAxis": {
            "type": "value",
            "name": record_names['local']['param'][info_name],
            "axisLabel": {
                'interval': 'auto',  # 根据图表的大小自动计算步长
            },
            'axisLine': {
                'show': True
            },
            'splitNumber': 5,  # 分成5个区间
        },
        'legend': {
            'data': ['客户' + str(cid) for cid in info_dict[mode_ref.value]],
            'type': 'scroll',  # 启用图例的滚动条
            'orient': 'horizontal',  # 横向排列
            'pageButtonItemGap': 5,
            'pageButtonGap': 20,
            'pageButtonPosition': 'end',  # 将翻页按钮放在最后
            'textStyle': {
                # 'width': 80,  # 设置图例文本的宽度
                'overflow': 'truncate',  # 当文本超出宽度时，截断文本
                'ellipsis': '...',  # 截断时末尾添加的字符串
            },
            'tooltip': {
                'show': True  # 启用悬停时的提示框
            }
        },
        'series': [
            {
                'name': '客户' + str(cid),
                'type': record_types['local']['param'][info_name],
                'data': list(info_dict[mode_ref.value][cid].value),  # 受制于rx，这里必须告知前端为list类型
                'connectNulls': True,  # 连接数据中的空值
            }
            for cid in info_dict[mode_ref.value]
        ],
        'dataZoom': [
            {
                'type': 'inside',  # 放大和缩小
                'orient': 'vertical',
                'start': 0,
                'end': 100,
                'minSpan': 1,  # 最小缩放比例，可以根据需要调整
                'maxSpan': 100,  # 最大缩放比例，可以根据需要调整
            },
            {
                'type': 'inside',
                'start': 0,
                'end': 100,
                'minSpan': 1,  # 最小缩放比例，可以根据需要调整
                'maxSpan': 100,  # 最大缩放比例，可以根据需要调整
            }
        ],
    }

# 全局信息使用算法-指标-轮次/时间的方式展示
def get_global_option_res(infos_dict, mode_ref, info_name, task_names):
    record_type = record_types['global']['param'][info_name]
    record_name = record_names['global']['param'][info_name]
    li = record_type.split('_')
    record_type = li[0]
    alloc_type = li[1] if len(li) > 1 else ''
    new_series, new_names, new_x = [], [], None
    if alloc_type == 'bul' and len(infos_dict[mode_ref.value][0]) != 0:
        new_names = list(list(infos_dict[mode_ref.value][0])[-1][1].keys())
        if record_type == 'line':  # 需要按x轴值排序
            paired_data = [
                (item[0], {name: item[1][name] for name in new_names})
                for item in infos_dict[mode_ref.value][0].value
            ]
            paired_data.sort(key=lambda x: x[0])
        else:
            paired_data = infos_dict[mode_ref.value][0]
        for name in new_names:
            data = [(item[0], item[1][name]) for item in paired_data]
            new_series.append({
                'name': name,
                'data': data,
                'type': record_type
            })
    options = {
        "toolbox": {
            "show": True,
            "feature": {
                "saveAsImage": {
                    "show": True,
                    "title": "下载图像",
                    "type": "png",
                    "lang": ["点击右键 -> 保存图片"]  # 提示用户如何保存图片
                }
            }
        },
        'grid': {
            "left": '10%',
            "top": '10%',
            "right": '10%',
            "bottom": '10%',
            "containLabel": True
        },
        'tooltip': {
            'trigger': 'axis',
            'axisPointer': {
                'type': 'cross',
                'lineStyle': {  # 设置纵向指示线
                    'type': 'dashed',
                    'color': "rgba(198, 196, 196, 0.75)"
                }
            },
            'crossStyle': {  # 设置横向指示线
                'color': "rgba(198, 196, 196, 0.75)"
            },
            'formatter': "算法{a}<br/>" + record_names['global']['type'][
                mode_ref.value] + ',' + record_type + "<br/>{c}",
            'extraCssText': 'box-shadow: 0 0 8px rgba(0, 0, 0, 0.3);'  # 添加阴影效果
        },
        "xAxis": {
            "type": 'category' if record_type == 'bar' else 'value',
            "name": record_names['global']['type'][mode_ref.value]
            if info_name not in type_name_mapping else type_name_mapping[info_name],
            'minInterval': 1 if mode_ref.value == 'round' else None,
        },
        "yAxis": {
            "type": "value",
            "name": record_name,
            "axisLabel": {
                'interval': 'auto',  # 根据图表的大小自动计算步长
            },
            'axisLine': {
                'show': True
            },
            'splitNumber': 5,  # 分成5个区间
        },
        'legend': {
            'data': [task_names[tid] for tid in task_names] if alloc_type != 'bul' or
                                                               len(infos_dict[mode_ref.value][
                                                                       0]) == 0 else new_names,
            'type': 'scroll',  # 启用图例的滚动条
            'orient': 'horizontal',  # 横向排列
            'pageButtonItemGap': 5,
            'pageButtonGap': 20,
            'pageButtonPosition': 'end',  # 将翻页按钮放在最后
            'textStyle': {
                'overflow': 'truncate',  # 当文本超出宽度时，截断文本
                'ellipsis': '...',  # 截断时末尾添加的字符串
            },
            'tooltip': {
                'show': True  # 启用悬停时的提示框
            }
        },
        'series': [
            {
                'name': task_names[tid],
                'type': record_type,
                'data': infos_dict[mode_ref.value][tid],  # 受制于rx，这里必须告知前端为list类型
                'connectNulls': True,  # 连接数据中的空值
            }
            for tid in infos_dict[mode_ref.value]
        ] if alloc_type != 'bul' or len(infos_dict[mode_ref.value][0]) == 0 else new_series,
        'dataZoom': [
            {
                'type': 'inside',  # 放大和缩小
                'orient': 'vertical',
                'start': 0,
                'end': 100,
                'minSpan': 1,  # 最小缩放比例，可以根据需要调整
                'maxSpan': 100,  # 最大缩放比例，可以根据需要调整
            },
            {
                'type': 'inside',
                'start': 0,
                'end': 100,
                'minSpan': 1,  # 最小缩放比例，可以根据需要调整
                'maxSpan': 100,  # 最大缩放比例，可以根据需要调整
            }
        ],
    }
    if alloc_type == 'seg' and infos_dict[mode_ref.value][0]:
        m = max(max(t) for t in infos_dict[mode_ref.value][0])
        options['series'].append({
            "type": 'line',
            "data": [[0, 0], [m, m]],  # y=x
            "lineStyle": {
                "color": 'red',
                "type": 'dashed',
                "width": 1
            }
        })
    return options

def get_local_option_res(info_dict: dict, mode_ref, info_name: str):
    return {
        "toolbox": {
            "show": True,
            "feature": {
                "saveAsImage": {
                    "show": True,
                    "title": "下载图像",
                    "type": "png",
                    "lang": ["点击右键 -> 保存图片"]  # 提示用户如何保存图片
                }
            }
        },
        'grid': {
            "left": '10%',
            "top": '10%',
            "right": '10%',
            "bottom": '10%',
            "containLabel": True
        },
        'tooltip': {
            'trigger': 'axis',
            'axisPointer': {
                'type': 'cross',
                'lineStyle': {  # 设置纵向指示线
                    'type': 'dashed',
                    'color': "rgba(198, 196, 196, 0.75)"
                }
            },
            'crossStyle': {  # 设置横向指示线
                'color': "rgba(198, 196, 196, 0.75)"
            },
            'formatter': "客户{a}<br/>" + record_names['local']['type'][mode_ref.value] + ',' +
                         record_names['local']['param'][info_name] + "<br/>{c}",
            'extraCssText': 'box-shadow: 0 0 8px rgba(0, 0, 0, 0.3);'  # 添加阴影效果
        },
        "xAxis": {
            "type": 'value',
            "name": record_names['local']['type'][mode_ref.value],
            'minInterval': 1 if mode_ref.value == 'round' else None,
        },
        "yAxis": {
            "type": "value",
            "name": record_names['local']['param'][info_name],
            "axisLabel": {
                'interval': 'auto',  # 根据图表的大小自动计算步长
            },
            'axisLine': {
                'show': True
            },
            'splitNumber': 5,  # 分成5个区间
        },
        'legend': {
            'data': ['客户' + str(cid) for cid in info_dict[mode_ref.value]],
            'type': 'scroll',  # 启用图例的滚动条
            'orient': 'horizontal',  # 横向排列
            'pageButtonItemGap': 5,
            'pageButtonGap': 20,
            'pageButtonPosition': 'end',  # 将翻页按钮放在最后
            'textStyle': {
                # 'width': 80,  # 设置图例文本的宽度
                'overflow': 'truncate',  # 当文本超出宽度时，截断文本
                'ellipsis': '...',  # 截断时末尾添加的字符串
            },
            'tooltip': {
                'show': True  # 启用悬停时的提示框
            }
        },
        'series': [
            {
                'name': '客户' + str(cid),
                'type': record_types['local']['param'][info_name],
                'data': info_dict[mode_ref.value][cid],  # 受制于rx，这里必须告知前端为list类型
                'connectNulls': True,  # 连接数据中的空值
            }
            for cid in info_dict[mode_ref.value]
        ],
        'dataZoom': [
            {
                'type': 'inside',  # 放大和缩小
                'orient': 'vertical',
                'start': 0,
                'end': 100,
                'minSpan': 1,  # 最小缩放比例，可以根据需要调整
                'maxSpan': 100,  # 最大缩放比例，可以根据需要调整
            },
            {
                'type': 'inside',
                'start': 0,
                'end': 100,
                'minSpan': 1,  # 最小缩放比例，可以根据需要调整
                'maxSpan': 100,  # 最大缩放比例，可以根据需要调整
            }
        ],
    }

def get_local_download_path(info_name, task_name=None, suffix='.xlsx'):
    """创建下载文件的名称，包含时间戳。"""
    current_time = datetime.now().strftime("%Y%m%d_%H%M%S")
    if task_name is None:
        filename = f"{info_name}_{current_time}{suffix}"
    else:
        filename = f"{task_name}_{info_name}_{current_time}{suffix}"
    print(filename)
    return filename

def transform_info_dicts(original_info):
    transformed_info = {}
    for data_type, tasks_data in original_info.items():
        for task_id, values in tasks_data.items():
            if task_id not in transformed_info:
                transformed_info[task_id] = {}
            transformed_info[task_id][data_type] = values  # 保留完整的[索引, 值]对
    return transformed_info

def download_global_info(info_name, task_names, info_dicts, mode_mapping, mode_ref):
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = 'All_Tasks'
    # 初始化行和列索引
    row_index = 2  # 第一行为任务名称，第二行开始为数据类型和数据
    column_index = 1
    record_name = record_names['global']['param'][info_name]
    record_type = record_types['global']['param'][info_name]
    li = record_type.split('_')
    record_type = li[0]
    alloc_type = li[1] if len(li) > 1 else ''
    new_series, new_names = [], []
    if alloc_type == 'bul' and len(info_dicts[mode_ref.value][0]) != 0:
        new_names = list(list(info_dicts[mode_ref.value][0])[-1][1].keys())
        if record_type == 'line':  # 需要按x轴值排序
            paired_data = [
                (item[0], {name: item[1][name] for name in new_names})
                for item in info_dicts[mode_ref.value][0].value
            ]
            paired_data.sort(key=lambda x: x[0])
        else:
            paired_data = info_dicts[mode_ref.value][0]
        for name in new_names:
            data = [(item[0], item[1][name]) for item in paired_data]
            new_series.append({
                'name': name,
                'data': data,
                'type': record_type
            })
        for item in new_series:
            t_name = item['name']
            data = item['data']

            # 合并任务名称所在的列，并居中对齐任务名称
            ws.merge_cells(start_row=1, start_column=column_index, end_row=1, end_column=column_index + 1)
            cell = ws.cell(row=1, column=column_index)
            cell.value = t_name
            cell.alignment = Alignment(horizontal='center')

            ws.cell(row=row_index, column=column_index).value = type_name_mapping[info_name]
            for i, pair in enumerate(data):
                ws.cell(row=row_index + 1 + i, column=column_index).value = pair[0]  # 从第三行开始写入数据
            column_index += 1

            # 在最后一个数据类型之后添加infoname标签并写入数据
            ws.cell(row=row_index, column=column_index).value = record_name
            for i, pair in enumerate(data):
                ws.cell(row=row_index + 1 + i, column=column_index).value = pair[1]  # 写入info_name对应的数据
            column_index += 2  # 跳过一列作为分隔列后，继续下一个任务
    else:
        info_dicts = transform_info_dicts(info_dicts)
        for tid, t_name in task_names.items():
            data_types = list(info_dicts[tid].keys())
            num_data_types = len(data_types)

            # 合并任务名称所在的列，并居中对齐任务名称
            ws.merge_cells(start_row=1, start_column=column_index, end_row=1, end_column=column_index + num_data_types)
            cell = ws.cell(row=1, column=column_index)
            cell.value = t_name
            cell.alignment = Alignment(horizontal='center')

            # 写入数据类型名称和对应数据
            for data_type in data_types:
                ws.cell(row=row_index, column=column_index).value = mode_mapping[data_type] \
                    if info_name not in type_name_mapping else type_name_mapping[info_name]
                for i, pair in enumerate(info_dicts[tid][data_type]):
                    ws.cell(row=row_index + 1 + i, column=column_index).value = pair[0]  # 从第三行开始写入数据
                column_index += 1

            # 在最后一个数据类型之后添加infoname标签并写入数据
            ws.cell(row=row_index, column=column_index).value = record_name
            for i, pair in enumerate(info_dicts[tid][data_types[0]]):
                ws.cell(row=row_index + 1 + i, column=column_index).value = pair[1]  # 写入info_name对应的数据

            column_index += 2  # 跳过一列作为分隔列后，继续下一个任务

    # 写入数据到内存中的字节流
    data = io.BytesIO()
    wb.save(data)
    data.seek(0)
    ui.download(data.read(), get_local_download_path(info_name))


def download_global_infos(exp_name, task_names, infos_dicts):
    # 创建工作簿和工作表
    wb = Workbook()
    wb.remove(wb.active)
    for info_name, info_dicts in infos_dicts.items():
        # 初始化行和列索引
        row_index = 2  # 第一行为任务名称，第二行开始为数据类型和数据
        column_index = 1
        record_name = record_names['global']['param'][info_name]
        ws = wb.create_sheet(title=record_name)
        record_type = record_types['global']['param'][info_name]
        li = record_type.split('_')
        record_type = li[0]
        alloc_type = li[1] if len(li) > 1 else ''
        new_series, new_names = [], []
        mode_list = list(info_dicts.keys())
        mode_mapping = {mode: record_names['global']['type'][mode] for mode in mode_list}
        if alloc_type == 'bul' and len(info_dicts[mode_list[0]][0]) != 0:
            new_names = list(list(info_dicts[mode_list[0]][0])[-1][1].keys())
            if record_type == 'line':  # 需要按x轴值排序
                paired_data = [
                    (item[0], {name: item[1][name] for name in new_names})
                    for item in info_dicts[mode_list[0]][0].value
                ]
                paired_data.sort(key=lambda x: x[0])
            else:
                paired_data = info_dicts[mode_list[0]][0]
            for name in new_names:
                data = [(item[0], item[1][name]) for item in paired_data]
                new_series.append({
                    'name': name,
                    'data': data,
                    'type': record_type
                })
            for item in new_series:
                t_name = item['name']
                data = item['data']

                # 合并任务名称所在的列，并居中对齐任务名称
                ws.merge_cells(start_row=1, start_column=column_index, end_row=1, end_column=column_index + 1)
                cell = ws.cell(row=1, column=column_index)
                cell.value = t_name
                cell.alignment = Alignment(horizontal='center')

                ws.cell(row=row_index, column=column_index).value = type_name_mapping[info_name]
                for i, pair in enumerate(data):
                    ws.cell(row=row_index + 1 + i, column=column_index).value = pair[0]  # 从第三行开始写入数据
                column_index += 1

                # 在最后一个数据类型之后添加infoname标签并写入数据
                ws.cell(row=row_index, column=column_index).value = record_name
                for i, pair in enumerate(data):
                    ws.cell(row=row_index + 1 + i, column=column_index).value = pair[1]  # 写入info_name对应的数据
                column_index += 2  # 跳过一列作为分隔列后，继续下一个任务
        else:
            info_dicts = transform_info_dicts(info_dicts)
            for tid, t_name in task_names.items():
                data_types = list(info_dicts[tid].keys())
                num_data_types = len(data_types)

                # 合并任务名称所在的列，并居中对齐任务名称
                ws.merge_cells(start_row=1, start_column=column_index, end_row=1, end_column=column_index + num_data_types)
                cell = ws.cell(row=1, column=column_index)
                cell.value = t_name
                cell.alignment = Alignment(horizontal='center')

                # 写入数据类型名称和对应数据
                for data_type in data_types:
                    ws.cell(row=row_index, column=column_index).value = mode_mapping[data_type] \
                        if info_name not in type_name_mapping else type_name_mapping[info_name]
                    for i, pair in enumerate(info_dicts[tid][data_type]):
                        ws.cell(row=row_index + 1 + i, column=column_index).value = pair[0]  # 从第三行开始写入数据
                    column_index += 1

                # 在最后一个数据类型之后添加infoname标签并写入数据
                ws.cell(row=row_index, column=column_index).value = record_name
                for i, pair in enumerate(info_dicts[tid][data_types[0]]):
                    ws.cell(row=row_index + 1 + i, column=column_index).value = pair[1]  # 写入info_name对应的数据

                column_index += 2  # 跳过一列作为分隔列后，继续下一个任务

    # 写入数据到内存中的字节流
    data = io.BytesIO()
    wb.save(data)
    data.seek(0)
    ui.download(data.read(), get_local_download_path('全局全局信息', exp_name))

def download_local_info(info_name, task_name, info_dicts, mode_mapping):
    record_name = record_names['local']['param'][info_name]
    info_dicts = transform_info_dicts(info_dicts)
    # 创建工作簿和工作表
    wb = Workbook()
    ws = wb.active
    ws.title = 'All_Clients'
    # 初始化行和列索引
    row_index = 2  # 第一行为任务名称，第二行开始为数据类型和数据
    column_index = 1
    for cid, value in info_dicts.items():
        data_types = list(value.keys())
        num_data_types = len(data_types)

        # 合并任务名称所在的列，并居中对齐任务名称
        ws.merge_cells(start_row=1, start_column=column_index, end_row=1, end_column=column_index + num_data_types)
        cell = ws.cell(row=1, column=column_index)
        cell.value = '客户' + str(cid)
        cell.alignment = Alignment(horizontal='center')

        # 写入数据类型名称和对应数据
        for data_type in data_types:
            ws.cell(row=row_index, column=column_index).value = mode_mapping[data_type] \
                if info_name not in type_name_mapping else type_name_mapping[info_name]
            for i, pair in enumerate(info_dicts[cid][data_type]):
                ws.cell(row=row_index + 1 + i, column=column_index).value = pair[0]  # 从第三行开始写入数据
            column_index += 1

        # 在最后一个数据类型之后添加infoname标签并写入数据
        ws.cell(row=row_index, column=column_index).value = record_name
        for i, pair in enumerate(info_dicts[cid][data_types[0]]):
            ws.cell(row=row_index + 1 + i, column=column_index).value = pair[1]  # 写入info_name对应的数据

        column_index += 2  # 跳过一列作为分隔列后，继续下一个任务
    # 写入数据到内存中的字节流
    data = io.BytesIO()
    wb.save(data)
    data.seek(0)
    ui.download(data.read(), get_local_download_path(info_name, task_name))

def download_local_infos(task_name, infos_dicts):
    # 创建工作簿和工作表
    wb = Workbook()
    wb.remove(wb.active)
    for info_name, info_dicts in infos_dicts.items():
        # 创建工作表并设置标题
        record_name = record_names['local']['param'][info_name]
        ws = wb.create_sheet(title=record_name)
        # 初始化行和列索引
        row_index = 2  # 第一行为任务名称，第二行开始为数据类型和数据
        column_index = 1

        info_dicts = transform_info_dicts(info_dicts)
        for cid, value in info_dicts.items():
            data_types = list(value.keys())
            num_data_types = len(data_types)

            # 合并任务名称所在的列，并居中对齐任务名称
            ws.merge_cells(start_row=1, start_column=column_index, end_row=1, end_column=column_index + num_data_types)
            cell = ws.cell(row=1, column=column_index)
            cell.value = '客户' + str(cid)
            cell.alignment = Alignment(horizontal='center')
            mode_mapping = {mode: record_names['local']['type'][mode] for mode in list(infos_dicts[info_name].keys())}
            # 写入数据类型名称和对应数据
            for data_type in data_types:
                ws.cell(row=row_index, column=column_index).value = mode_mapping[data_type]
                for i, pair in enumerate(info_dicts[cid][data_type]):
                    ws.cell(row=row_index + 1 + i, column=column_index).value = pair[0]  # 从第三行开始写入数据
                column_index += 1

            # 在最后一个数据类型之后添加infoname标签并写入数据
            ws.cell(row=row_index, column=column_index).value = record_name
            for i, pair in enumerate(info_dicts[cid][data_types[0]]):
                ws.cell(row=row_index + 1 + i, column=column_index).value = pair[1]  # 写入info_name对应的数据

            column_index += 2  # 跳过一列作为分隔列后，继续下一个任务
    # 写入数据到内存中的字节流
    data = io.BytesIO()
    wb.save(data)
    data.seek(0)
    ui.download(data.read(), get_local_download_path('全部本地信息', task_name))



def control_global_echarts(info_name, infos_dicts, task_names, is_res=False):
    mode_list = list(infos_dicts.keys())
    mode_mapping = {mode: record_names['global']['type'][mode] for mode in mode_list}
    mode_ref = to_ref(mode_list[0])
    with ui.column().classes("h-[40rem] w-[40rem]"):
        with ui.row():
            rxui.select(value=mode_ref, options=mode_mapping)
            if is_res:
                rxui.button('下载数据', on_click=lambda: \
                    download_global_info(info_name, task_names, infos_dicts, mode_mapping, mode_ref)).props('icon=cloud_download')
        if is_res:
            rxui.echarts(lambda: get_global_option_res(infos_dicts, mode_ref, info_name, task_names),
                         not_merge=False).classes("h-[40rem] w-[40rem]")
        else:
            rxui.echarts(lambda: get_global_option(infos_dicts, mode_ref, info_name, task_names),
                         not_merge=False).classes("h-[40rem] w-[40rem]")


def control_local_echarts(infos_dicts, is_res=False, task_name=None):
    with rxui.grid(columns=2).classes('w-full h-full'):

        for info_name in infos_dicts:
            with ui.column().classes("h-[40rem] w-[40rem]"):
                mode_list = list(infos_dicts[info_name].keys())
                mode_ref = to_ref(mode_list[0])
                mode_mapping = {mode: record_names['local']['type'][mode] for mode in mode_list}
                with ui.row():
                    rxui.select(value=mode_ref, options=mode_mapping)
                    if is_res:
                        rxui.button('下载数据', on_click=lambda: download_local_info(info_name, task_name if task_name is not None else '暂无任务名称',
                                                                                     infos_dicts[info_name], mode_mapping)).props('icon=cloud_download')
                if is_res:
                    rxui.echarts(
                        lambda mode_ref=mode_ref, info_name=info_name: get_local_option_res(infos_dicts[info_name],
                                                                                             mode_ref, info_name),
                        not_merge=False).classes("h-[40rem] w-[40rem]")
                else:
                    rxui.echarts(
                        lambda mode_ref=mode_ref, info_name=info_name: get_local_option(infos_dicts[info_name], mode_ref,
                                                                                        info_name),
                        not_merge=False).classes("h-[40rem] w-[40rem]")


async def move_all_files(old_path, new_path):
    try:
        os.makedirs(new_path, exist_ok=True)  # 确保新目录存在
        for filename in os.listdir(old_path):
            old_file = os.path.join(old_path, filename)
            new_file = os.path.join(new_path, filename)
            if os.path.isfile(old_file):  # 确保是文件而不是目录
                shutil.move(old_file, new_file)
        print(f"所有文件已从{old_path}移动到{new_path}")
    except Exception as e:
        print(f"移动文件时出错: {e}")


def locked_page_height():
    """
    锁定页面内容区域的高度，等于屏幕高度减去顶部和底部的高度
    意味着 内容区创建的第一个容器，可以通过 h-full 让其高度等于屏幕高度(减去顶部和底部的高度).

    此函数创建时的 nicegui 版本为:1.4.20
    """
    client = context.client
    q_page = client.page_container.default_slot.children[0]
    q_page.props(
        ''':style-fn="(offset, height) => ( { height: offset ? `calc(100vh - ${offset}px)` : '100vh' })"'''
    )
    client.content.classes("h-full")


async def test_api_access(api_key, api_base, model=None, model_type="generation") -> (bool, str):
    """
    异步测试API key和API基地址能否访问特定模型，并记录测试时间。

    参数:
    - api_key: API密钥。
    - api_base: API基地址。
    - model: 要测试的模型名称。如果未提供，则使用默认模型进行测试。

    返回值:
    - bool: 表示是否成功访问。
    - str: 访问成功或失败的详细信息，包括测试时间（用中文回显）。
    """
    if model is None:
        model = "gpt-3.5-turbo"

    async with aiohttp.ClientSession() as session:
        url = f"{api_base}/completions"
        headers = {
            "Authorization": f"Bearer {api_key}",
            "Content-Type": "application/json"
        }
        if model_type == "generation":
            if model is None:
                model = "gpt-3.5-turbo"
            url = f"{api_base}/completions"
            data = {
                "model": model,
                "prompt": "Hello, world!",
                "max_tokens": 5
            }
        elif model_type == "embedding":
            if model is None:
                model = "text-embedding-ada-002"
            url = f"{api_base}/embeddings"
            data = {
                "model": model,
                "input": "Hello, world!"
            }
        else:
            return False, "未知的模型类型。"

        start_time = time.perf_counter()
        try:
            async with session.post(url, headers=headers, json=data) as response:
                elapsed = time.perf_counter() - start_time
                status = response.status
                try:
                    response_body = await response.json()
                except Exception:
                    response_body = await response.text()

                if status == 200:
                    return True, f"访问成功。测试时间：{elapsed:.2f}秒。"
                elif status == 401:
                    return False, f"未授权访问。请检查API密钥是否正确。测试时间：{elapsed:.2f}秒。"
                elif status == 403:
                    return False, f"禁止访问。API密钥缺少权限。测试时间：{elapsed:.2f}秒。"
                elif status == 404:
                    return False, f"资源未找到。请检查API基地址或模型名称。测试时间：{elapsed:.2f}秒。"
                elif status == 429:
                    return False, f"请求过多。已超出速率限制。测试时间：{elapsed:.2f}秒。"
                else:
                    return False, f"发生错误：HTTP {status}。响应：{response_body}。测试时间：{elapsed:.2f}秒。"

        except aiohttp.ClientConnectorError as e:
            return False, f"无法连接到主机。可能是网络问题、目标服务器不可达、或SSL问题。异常信息：{e}"
        except Exception as e:
            return False, f"发生未预期的错误。异常信息：{e}"


def clear_ref(info_dict, name=None):
    for k, v in info_dict.items():
        if type(v) == dict:
            if name is None or name == k:
                clear_ref(v)
            else:
                clear_ref(v, name)
        elif name is None or name == k:
            v.value.clear()


def read_and_modify_svg(file_path, new_gradient):
    # 读取 SVG 文件
    with open(file_path, 'r', encoding='utf-8') as file:
        svg_content = file.read()
    # 替换现有的渐变定义
    # 假设原SVG中有一个渐变定义可以被识别为 '<linearGradient...'
    # 这里需要确保这个标记与你的SVG文件中的实际内容匹配
    import re
    new_content = re.sub(r'<linearGradient.*?</linearGradient>',
                         new_gradient, svg_content, flags=re.DOTALL)
    return new_content